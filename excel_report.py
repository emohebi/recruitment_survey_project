"""
Excel report generator: writes formatted multi-sheet workbook.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_HEADER_COLOR, EXCEL_FONT, BREAKDOWNS


PCT_COLS = {"RR", "RD", "FSI", "Cost_Concern"}

HEADER_FONT = Font(name=EXCEL_FONT, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=EXCEL_HEADER_COLOR)
CELL_FONT = Font(name=EXCEL_FONT, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _write_sheet(ws, df: pd.DataFrame):
    """Write a DataFrame to a worksheet with formatting."""
    headers = list(df.columns)

    # Header row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.font = CELL_FONT
            col_name = headers[ci - 1]
            if col_name in PCT_COLS and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
            elif col_name == "Pct" and isinstance(val, (int, float)):
                cell.number_format = "0.0"
            elif isinstance(val, (int, float)) and col_name not in ("N",):
                cell.number_format = "#,##0.00"

    # Column widths
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(15, len(str(headers[ci - 1])) + 4)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _col_order(df: pd.DataFrame, lead_cols: list) -> pd.DataFrame:
    """Reorder columns: lead_cols first, then indicators, then the rest."""
    indicator_cols = [c for c in ["RR", "RD", "FSI", "Cost_Concern"] if c in df.columns]
    rest = [c for c in df.columns if c not in lead_cols and c not in indicator_cols]
    ordered = [c for c in lead_cols if c in df.columns] + indicator_cols + rest
    return df[ordered]


def write_excel(results: dict, output_path: str):
    """
    Write all analysis results to a formatted Excel workbook.

    Parameters
    ----------
    results : dict of DataFrames from indicators.run_all_analyses()
    output_path : file path for the .xlsx output
    """
    wb = Workbook()

    # 1. National
    ws = wb.active
    ws.title = "National"
    _write_sheet(ws, _col_order(results["national"], ["Week"]))

    # 2. Breakdown sheets
    sheet_specs = [
        ("by_State",        "By State",             ["Week", "State"]),
        ("by_Industry",     "By Industry",          ["Week", "Industry06"]),
        ("by_GCC",          "By Region (GCC)",      ["Week", "GCC"]),
        ("by_BusinessSize", "By Business Size",     ["Week", "BusinessSize"]),
        ("by_ARIA",         "By Remoteness (ARIA)", ["Week", "ARIA"]),
        ("by_CCRoS",        "By Capital-Rest",      ["Week", "CCRoS"]),
    ]
    for key, title, leads in sheet_specs:
        if key in results and not results[key].empty:
            ws = wb.create_sheet(title)
            _write_sheet(ws, _col_order(results[key], leads))

    # 3. Cost impact sheets
    cost_specs = [
        ("cost_State",    "Cost Impact - State",    ["State"]),
        ("cost_Industry", "Cost Impact - Industry", ["Industry06"]),
        ("cost_GCC",      "Cost Impact - Region",   ["GCC"]),
    ]
    for key, title, leads in cost_specs:
        if key in results and not results[key].empty:
            ws = wb.create_sheet(title)
            _write_sheet(ws, results[key])

    # 4. Concern distribution
    if "concern_distribution" in results and not results["concern_distribution"].empty:
        ws = wb.create_sheet("Concern Distribution")
        _write_sheet(ws, results["concern_distribution"])

    wb.save(output_path)
    print(f"Excel report saved → {output_path}")
